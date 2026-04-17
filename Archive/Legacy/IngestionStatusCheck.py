#!/usr/bin/env python3
"""Find URLs from an input list that do not exist in a content database export.

Why this helps:
- Handles common URL formatting differences (scheme/case/trailing slash/query order)
- Optionally follows redirects for input URLs before matching
- Optionally matches GUIDs extracted from URLs against a DB GUID column

Usage examples:
    python IngestionStatusCheck.py \
    --input missing_urls.txt \
    --db content_export.csv \
    --db-url-column canonical_url \
    --db-guid-column content_id \
    --output truly_missing.csv

    python IngestionStatusCheck.py \
    --input incoming_urls.csv --input-column url \
    --db db_urls.csv --db-url-column url \
    --no-resolve-redirects

    python IngestionStatusCheck.py \
        --input input_urls.xlsx --input-sheet Sheet1 \
        --db db_urls.xlsx --db-sheet Sheet1
"""

from __future__ import annotations

import argparse
import base64
import csv
import os
import re
import subprocess
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlparse, urlunparse
from urllib.request import Request, urlopen

DEFAULT_INGESTED_SOURCE = (
    "https://microsoft.sharepoint.com/:x:/t/SxG/"
    "IQBf9XzR8TvmQ5DkGe_tSu03ATBm-KUwtYMAPgQxLFyBJDM?e=PlaIVb"
)
DEFAULT_BLOCKED_SOURCE = (
    "https://microsoft.sharepoint.com/:x:/t/SxG/"
    "IQDo8noAWQFJQr7v_6GapyESATbH9L5jZcotqMFPcu403x8?e=waSU1k"
)

FALLBACK_SOURCE_FILES = {
    "ingested": "IngestedURLs.xlsx",
    "blocked": "BlockedURLs.xlsx",
}

GUID_RE = re.compile(
    r"[0-9a-fA-F]{8}-"
    r"[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{12}"
)

TRACKING_QUERY_PREFIXES = (
    "utm_",
    "mc_",
)

TRACKING_QUERY_KEYS = {
    "gclid",
    "fbclid",
    "msclkid",
    "igshid",
    "yclid",
    "_hsenc",
    "_hsmi",
}

COMMON_URL_COLUMNS = (
    "url",
    "uri",
    "link",
    "href",
    "canonical_url",
    "slug_url",
)


@dataclass
class ResolveResult:
    original_url: str
    resolved_url: str | None
    error: str | None


@dataclass
class MatchResult:
    input_url: str
    resolved_url: str | None
    status: str
    match_reason: str
    matched_value: str
    guids_found: str
    resolve_error: str


def classify_match(
    input_url: str,
    resolved_url: str | None,
    resolve_error: str | None,
    db_full: set[str],
    db_path: set[str],
    db_guid: set[str],
    blocked_full: set[str] | None = None,
    blocked_path: set[str] | None = None,
) -> MatchResult:
    input_full, input_path, input_guids = get_keys_for_input(input_url, resolved_url)
    guids_str = ",".join(sorted(input_guids))

    is_login_required = is_login_required_redirect(input_url, resolved_url, resolve_error)

    full_hit = next((k for k in input_full if k in db_full), None)
    if full_hit:
        # Found in DB — check whether it is also blocked.
        if blocked_full:
            blocked_hit = next((k for k in input_full if k in blocked_full), None)
            if blocked_hit:
                return MatchResult(input_url, resolved_url, "blocked", "blocked", blocked_hit, guids_str, resolve_error or "")
        return MatchResult(input_url, resolved_url, "found", "full-url", full_hit, guids_str, resolve_error or "")

    guid_hit = next((g for g in input_guids if g in db_guid), None)
    if guid_hit:
        return MatchResult(input_url, resolved_url, "found", "guid", guid_hit, guids_str, resolve_error or "")

    # No match found in DB. Only mark as not-testable if login is required.
    if is_login_required:
        return MatchResult(input_url, resolved_url, "not-testable", "login-required", "", guids_str, resolve_error or "")

    return MatchResult(input_url, resolved_url, "missing", "no-match", "", guids_str, resolve_error or "")


def normalize_url(raw_url: str) -> str:
    """Return a normalized URL string to improve match reliability."""
    value = (raw_url or "").strip()
    if not value:
        return ""

    parsed = urlparse(value)
    if not parsed.scheme or not parsed.netloc:
        return ""

    scheme = parsed.scheme.lower()
    host = parsed.hostname.lower() if parsed.hostname else ""

    port = parsed.port
    if port and not ((scheme == "http" and port == 80) or (scheme == "https" and port == 443)):
        host = f"{host}:{port}"

    path = parsed.path or "/"
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]

    cleaned_query_items = []
    for k, v in parse_qsl(parsed.query, keep_blank_values=True):
        k_l = k.lower()
        if k_l.startswith(TRACKING_QUERY_PREFIXES):
            continue
        if k_l in TRACKING_QUERY_KEYS:
            continue
        cleaned_query_items.append((k, v))

    cleaned_query_items.sort(key=lambda i: (i[0].lower(), i[1]))
    query = "&".join(f"{k}={v}" if v != "" else k for k, v in cleaned_query_items)

    return urlunparse((scheme, host, path, "", query, ""))


def is_login_required_redirect(input_url: str, resolved_url: str | None, resolve_error: str | None) -> bool:
    """Detect redirects/errors indicating the target requires authentication.

    This prevents auth-gated URLs (for example Azure DevOps wiki pages) from being
    mislabeled as found/missing when content cannot be validated anonymously.
    """
    input_host = (urlparse(input_url).hostname or "").lower()

    if resolved_url:
        parsed = urlparse(resolved_url)
        host = (parsed.hostname or "").lower()
        path = parsed.path.lower()
        query_map = {k.lower(): v.lower() for k, v in parse_qsl(parsed.query, keep_blank_values=True)}

        if (
            ("signin" in path or "_signin" in path)
            and (
                "visualstudio.com" in host
                or "microsoftonline.com" in host
                or "login.live.com" in host
            )
        ):
            return True

        if query_map.get("realm") == "dev.azure.com":
            return True

        if "dev.azure.com" in query_map.get("reply_to", ""):
            return True

    if resolve_error:
        err = resolve_error.lower()
        if (
            ("http error 401" in err or "http error 403" in err)
            and ("dev.azure.com" in input_host or "visualstudio.com" in input_host)
        ):
            return True

    return False


def no_query_key(normalized_url: str) -> str:
    if not normalized_url:
        return ""
    p = urlparse(normalized_url)
    return urlunparse((p.scheme, p.netloc, p.path, "", "", ""))


def extract_guids(text: str) -> set[str]:
    return {m.group(0).lower() for m in GUID_RE.finditer(text or "")}


def detect_column(fieldnames: list[str], preferred: str | None, candidates: Iterable[str]) -> str:
    if not fieldnames:
        raise ValueError("CSV has no header row")

    # If a CSV has exactly one column, treat it as the URL column automatically.
    if len(fieldnames) == 1 and not preferred:
        return fieldnames[0]

    if preferred:
        if preferred not in fieldnames:
            raise ValueError(f"Column '{preferred}' was not found. Available columns: {fieldnames}")
        return preferred

    lower_map = {f.lower(): f for f in fieldnames}
    for candidate in candidates:
        if candidate.lower() in lower_map:
            return lower_map[candidate.lower()]

    raise ValueError(
        "No matching column found. Provide --input-column/--db-url-column explicitly "
        f"or use a single-column CSV. Available: {fieldnames}"
    )


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as ex:  # pragma: no cover
        raise RuntimeError(
            "Reading .xlsx files requires openpyxl. Install with: pip install openpyxl"
        ) from ex
    return openpyxl


def _read_excel_rows(path: Path, sheet_name: str | None) -> tuple[list[str], list[dict[str, str]]]:
    openpyxl = _require_openpyxl()

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {workbook.sheetnames}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)

    header_row = None
    for candidate in rows_iter:
        if any(cell is not None and str(cell).strip() != "" for cell in candidate):
            header_row = candidate
            break

    if header_row is None:
        workbook.close()
        raise ValueError(f"Sheet '{sheet.title}' in {path} is empty")

    fieldnames = [(str(cell).strip() if cell is not None else "") for cell in header_row]
    if not any(fieldnames):
        workbook.close()
        raise ValueError(f"Sheet '{sheet.title}' in {path} does not contain a valid header row")

    # Fill unnamed columns with deterministic names so row mapping is stable.
    for i, name in enumerate(fieldnames):
        if not name:
            fieldnames[i] = f"column_{i + 1}"

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        row_values = list(raw_row)
        if len(row_values) < len(fieldnames):
            row_values.extend([None] * (len(fieldnames) - len(row_values)))
        row_map = {
            fieldnames[i]: ("" if row_values[i] is None else str(row_values[i]).strip())
            for i in range(len(fieldnames))
        }
        rows.append(row_map)

    workbook.close()
    return fieldnames, rows


def is_http_url(value: str) -> bool:
    parsed = urlparse((value or "").strip())
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def build_sharepoint_download_url(url: str) -> str:
    """Force direct download mode for SharePoint links when possible."""
    parsed = urlparse(url)
    if "sharepoint.com" not in (parsed.netloc or "").lower():
        return url

    query_items = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k.lower() != "e"]
    query_map = {k: v for k, v in query_items}
    query_map["download"] = "1"

    rebuilt_query = "&".join(f"{k}={v}" if v != "" else k for k, v in sorted(query_map.items(), key=lambda x: x[0].lower()))
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", rebuilt_query, ""))


def encode_share_url_for_graph(url: str) -> str:
    encoded = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8").rstrip("=")
    return f"u!{encoded}"


def get_graph_token_from_az() -> str:
    """Get a Microsoft Graph access token from Azure CLI login context."""
    proc = subprocess.run(
        [
            "az",
            "account",
            "get-access-token",
            "--resource-type",
            "ms-graph",
            "--query",
            "accessToken",
            "-o",
            "tsv",
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        stderr = (proc.stderr or "").strip()
        raise RuntimeError(
            "Unable to get Graph token from Azure CLI. Run `az login` and try again. "
            f"Details: {stderr or 'unknown error'}"
        )

    token = (proc.stdout or "").strip()
    if not token:
        raise RuntimeError("Azure CLI returned an empty Graph token. Run `az login` and try again.")
    return token


def get_graph_token_from_az_powershell() -> str:
    """Get a Microsoft Graph token from Az PowerShell (Connect-AzAccount context)."""
    proc = subprocess.run(
        ["powershell", "-NoProfile", "-Command", "(Get-AzAccessToken -ResourceTypeName MSGraph).Token"],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        stderr = (proc.stderr or "").strip()
        raise RuntimeError(
            "Unable to get Graph token from Az PowerShell. Run `Connect-AzAccount` and try again. "
            f"Details: {stderr or 'unknown error'}"
        )

    token = (proc.stdout or "").strip()
    if not token:
        raise RuntimeError("Az PowerShell returned an empty Graph token. Run `Connect-AzAccount` and try again.")
    return token


def download_sharepoint_via_graph(url: str) -> bytes:
    token = get_graph_token_from_az()
    share_id = encode_share_url_for_graph(url)
    graph_url = f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem/content"
    request = Request(graph_url, headers={"Authorization": f"Bearer {token}", "User-Agent": "ingestion-link-audit/1.0"})
    with urlopen(request, timeout=60) as response:
        return response.read()


def download_sharepoint_via_graph_az_powershell(url: str) -> bytes:
    token = get_graph_token_from_az_powershell()
    share_id = encode_share_url_for_graph(url)
    graph_url = f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem/content"
    request = Request(graph_url, headers={"Authorization": f"Bearer {token}", "User-Agent": "ingestion-link-audit/1.0"})
    with urlopen(request, timeout=60) as response:
        return response.read()


def download_sharepoint_via_powershell(url: str) -> bytes:
    """Use Windows integrated auth context to download SharePoint content."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_path = Path(temp_file.name)

    ps_command = (
        f"Invoke-WebRequest -Uri '{url}' -OutFile '{temp_path}'"
    )

    try:
        proc = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_command],
            capture_output=True,
            text=True,
            check=False,
        )
        if proc.returncode != 0:
            stderr = (proc.stderr or "").strip()
            raise RuntimeError(stderr or "PowerShell download failed")

        data = temp_path.read_bytes()
        if not data:
            raise RuntimeError("Downloaded file is empty")
        return data
    finally:
        try:
            if temp_path.exists():
                temp_path.unlink()
        except Exception:
            pass


def looks_like_html(data: bytes) -> bool:
    sample = data[:1024].decode("utf-8", errors="ignore").lower()
    return "<html" in sample or "<!doctype html" in sample or "sign in to your account" in sample


def looks_like_xlsx(data: bytes) -> bool:
    return data.startswith(b"PK\x03\x04") or data.startswith(b"PK\x05\x06") or data.startswith(b"PK\x07\x08")


def validate_downloaded_content(url: str, data: bytes, label: str) -> None:
    if not data:
        raise RuntimeError(f"Downloaded file for {label} is empty")

    if looks_like_html(data):
        raise RuntimeError("Downloaded content is an HTML sign-in page, not the data file")

    expect_xlsx = "/:x:/" in url.lower() or url.lower().endswith(".xlsx")
    if expect_xlsx and not looks_like_xlsx(data):
        raise RuntimeError("Downloaded content does not look like a valid XLSX file")


def find_local_fallback_file(label: str) -> Path | None:
    filename = FALLBACK_SOURCE_FILES.get(label)
    if not filename:
        return None

    home = Path.home()

    # When packaged as a PyInstaller EXE, sys.executable is the EXE itself.
    # __file__ resolves to the temp extraction dir, not the EXE folder.
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
    else:
        exe_dir = Path(__file__).resolve().parent

    # Build candidate paths — EXE folder first so side-by-side files win,
    # then common OneDrive sync locations for auto-updating SharePoint libraries.
    candidates: list[Path] = [
        exe_dir / filename,
        Path.cwd() / filename,
    ]

    # OneDrive for Business syncs SharePoint libraries to
    # ~/Microsoft/<SiteName> - <LibraryName>/  (Windows OneDrive sync client)
    # Glob all synced Microsoft site folders so any SharePoint library name works.
    microsoft_dir = home / "Microsoft"
    if microsoft_dir.is_dir():
        for site_dir in sorted(microsoft_dir.iterdir()):
            if site_dir.is_dir():
                candidates.append(site_dir / filename)

    # Explicit well-known paths as final fallbacks.
    candidates += [
        home / "OneDrive - Microsoft" / "MCfS Team Stuff" / "Desktop" / filename,
        home / "OneDrive - Microsoft" / filename,
        home / "Desktop" / filename,
    ]

    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def ensure_local_source(source: str, cache_dir: Path, label: str) -> Path:
    """Return a local file path for either a local source path or a downloadable URL."""
    value = (source or "").strip()
    if not value:
        raise ValueError(f"Missing source for {label}")

    if not is_http_url(value):
        return Path(value)

    # In enterprise environments SharePoint auth can be blocked. Prefer trusted
    # local exports first when available.
    prefer_local = os.getenv("URL_AUDIT_PREFER_LOCAL", "1").strip().lower() not in {"0", "false", "no"}
    if prefer_local:
        local_fallback = find_local_fallback_file(label)
        if local_fallback:
            print(f"Using local {label} source: {local_fallback}")
            return local_fallback

    cache_dir.mkdir(parents=True, exist_ok=True)
    local_path = cache_dir / f"{label}.xlsx"
    download_url = build_sharepoint_download_url(value)
    request = Request(download_url, headers={"User-Agent": "ingestion-link-audit/1.0"})

    try:
        with urlopen(request, timeout=60) as response:
            data = response.read()
        validate_downloaded_content(value, data, label)
    except Exception as ex:  # noqa: BLE001
        # For Microsoft-internal links, try Windows-integrated PowerShell download first.
        try:
            data = download_sharepoint_via_powershell(download_url)
            validate_downloaded_content(value, data, label)
        except Exception as ps_ex:  # noqa: BLE001
            # Next fallback: Graph download using Azure CLI auth.
            try:
                data = download_sharepoint_via_graph(value)
                validate_downloaded_content(value, data, label)
            except Exception as graph_az_ex:  # noqa: BLE001
                # Final fallback: Graph download using Az PowerShell auth.
                try:
                    data = download_sharepoint_via_graph_az_powershell(value)
                    validate_downloaded_content(value, data, label)
                except Exception as graph_ps_ex:  # noqa: BLE001
                    local_fallback = find_local_fallback_file(label)
                    if local_fallback:
                        print(
                            f"Warning: using local fallback for {label}: {local_fallback} "
                            "(SharePoint download/auth not available in this environment)"
                        )
                        return local_fallback
                    raise RuntimeError(
                        f"Failed to download {label} from URL. URL: {value}. "
                        f"Direct download failed with {type(ex).__name__}: {ex}. "
                        f"PowerShell fallback failed with {type(ps_ex).__name__}: {ps_ex}. "
                        f"Graph(Azure CLI) fallback failed with {type(graph_az_ex).__name__}: {graph_az_ex}. "
                        f"Graph(Az PowerShell) fallback failed with {type(graph_ps_ex).__name__}: {graph_ps_ex}"
                    ) from graph_ps_ex

    local_path.write_bytes(data)
    return local_path


def read_urls(
    path: Path,
    column: str | None,
    candidate_columns: Iterable[str],
    sheet_name: str | None = None,
    no_header: bool = False,
) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if path.suffix.lower() in {".txt", ".list"}:
        return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]

    if path.suffix.lower() == ".csv":
        if no_header:
            rows = []
            with path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if not row:
                        continue
                    value = (row[0] or "").strip()
                    if value:
                        rows.append(value)
            return rows

        rows = []
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            selected = detect_column(reader.fieldnames or [], column, candidate_columns)
            for row in reader:
                v = (row.get(selected) or "").strip()
                if v:
                    rows.append(v)
        return rows

    if path.suffix.lower() == ".xlsx":
        if no_header:
            openpyxl = _require_openpyxl()
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {workbook.sheetnames}")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            urls: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                if not row:
                    continue
                value = ("" if row[0] is None else str(row[0]).strip())
                if value:
                    urls.append(value)

            workbook.close()
            return urls

        fieldnames, rows = _read_excel_rows(path, sheet_name)
        selected = detect_column(fieldnames, column, candidate_columns)
        return [value for row in rows if (value := (row.get(selected) or "").strip())]

    raise ValueError(f"Unsupported file type for URL input: {path.suffix}. Use .txt, .csv, or .xlsx")


def read_db_guid_values(
    path: Path,
    guid_column: str | None,
    sheet_name: str | None = None,
    no_header: bool = False,
) -> set[str]:
    if not guid_column:
        return set()

    if no_header:
        raise ValueError("--db-guid-column cannot be used together with --db-no-header")

    results: set[str] = set()

    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            selected = detect_column(reader.fieldnames or [], guid_column, (guid_column,))
            for row in reader:
                value = (row.get(selected) or "").strip().lower()
                if value:
                    results.add(value)
        return results

    if path.suffix.lower() == ".xlsx":
        fieldnames, rows = _read_excel_rows(path, sheet_name)
        selected = detect_column(fieldnames, guid_column, (guid_column,))
        for row in rows:
            value = (row.get(selected) or "").strip().lower()
            if value:
                results.add(value)
        return results

    raise ValueError("--db-guid-column requires --db to be .csv or .xlsx")

    return results


def resolve_final_url(url: str, timeout: int) -> ResolveResult:
    if not url:
        return ResolveResult(original_url=url, resolved_url=None, error="empty-url")

    for method in ("HEAD", "GET"):
        try:
            req = Request(url, method=method, headers={"User-Agent": "ingestion-link-audit/1.0"})
            with urlopen(req, timeout=timeout) as response:
                final_url = response.geturl()
            return ResolveResult(original_url=url, resolved_url=final_url, error=None)
        except Exception as ex:  # noqa: BLE001
            last_err = f"{method}: {type(ex).__name__}: {ex}"

    return ResolveResult(original_url=url, resolved_url=None, error=last_err)


def build_db_keys(db_urls: list[str]) -> tuple[set[str], set[str], set[str]]:
    full_keys: set[str] = set()
    path_keys: set[str] = set()
    guid_keys: set[str] = set()

    for raw in db_urls:
        norm = normalize_url(raw)
        if norm:
            full_keys.add(norm)
            path_keys.add(no_query_key(norm))
            guid_keys.update(extract_guids(norm))
    return full_keys, path_keys, guid_keys


def get_keys_for_input(raw_url: str, resolved_url: str | None) -> tuple[set[str], set[str], set[str]]:
    urls_to_check = [raw_url]
    if resolved_url:
        urls_to_check.append(resolved_url)

    full_keys: set[str] = set()
    path_keys: set[str] = set()
    guid_keys: set[str] = set()

    for url in urls_to_check:
        norm = normalize_url(url)
        if norm:
            full_keys.add(norm)
            path_keys.add(no_query_key(norm))
        guid_keys.update(extract_guids(url))
        guid_keys.update(extract_guids(norm)) if norm else None

    return full_keys, path_keys, guid_keys


def classify_all_inputs(
    input_urls: list[str],
    db_full: set[str],
    db_path: set[str],
    db_guid: set[str],
    resolve_redirects: bool,
    timeout: int,
    workers: int,
    blocked_full: set[str] | None = None,
    blocked_path: set[str] | None = None,
) -> list[MatchResult]:
    resolve_map: dict[str, ResolveResult] = {}

    if resolve_redirects:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(resolve_final_url, url, timeout): url for url in input_urls}
            for future in as_completed(futures):
                url = futures[future]
                try:
                    resolve_map[url] = future.result()
                except Exception as ex:  # noqa: BLE001
                    resolve_map[url] = ResolveResult(url, None, f"internal-error: {type(ex).__name__}: {ex}")
    else:
        for u in input_urls:
            resolve_map[u] = ResolveResult(u, None, None)

    results: list[MatchResult] = []

    for input_url in input_urls:
        resolved = resolve_map[input_url]
        results.append(
            classify_match(
                input_url=input_url,
                resolved_url=resolved.resolved_url,
                resolve_error=resolved.error,
                db_full=db_full,
                db_path=db_path,
                db_guid=db_guid,
                blocked_full=blocked_full,
                blocked_path=blocked_path,
            )
        )

    return results


def write_results(output_path: Path, rows: list[MatchResult]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "input_url",
                "resolved_url",
                "status",
                "match_reason",
                "matched_value",
                "guids_found",
                "resolve_error",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.input_url,
                    row.resolved_url or "",
                    row.status,
                    row.match_reason,
                    row.matched_value,
                    row.guids_found,
                    row.resolve_error,
                ]
            )


def pick_input_file_interactively() -> Path:
    """Prompt the user to pick an input file when --input is not provided."""
    filetypes = [
        ("Supported files", "*.xlsx *.csv *.txt *.list"),
        ("Excel files", "*.xlsx"),
        ("CSV files", "*.csv"),
        ("Text files", "*.txt *.list"),
        ("All files", "*.*"),
    ]

    # Prefer a native picker for non-technical users when GUI is available.
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        selected = filedialog.askopenfilename(title="Select URL input file", filetypes=filetypes)
        root.destroy()
        if selected:
            return Path(selected)
    except Exception:
        pass

    entered = input("Enter path to input URL file (.xlsx/.csv/.txt): ").strip().strip('"')
    if not entered:
        raise ValueError("No input file selected")
    return Path(entered)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Find input URLs truly missing from DB URL export")
    parser.add_argument("--input", required=False, type=Path, help="Input URL list (.txt, .csv, or .xlsx)")
    parser.add_argument("--input-column", default=None, help="Column name for input URLs when --input is CSV/XLSX")
    parser.add_argument("--input-sheet", default=None, help="Worksheet name when --input is .xlsx (default: active)")
    parser.add_argument(
        "--input-no-header",
        action="store_true",
        help="Treat input CSV/XLSX as headerless and read URLs from first column",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_INGESTED_SOURCE,
        help="Database export source: local path or URL (.txt, .csv, .xlsx)",
    )
    parser.add_argument("--db-url-column", default=None, help="DB URL column when --db is CSV/XLSX")
    parser.add_argument("--db-sheet", default=None, help="Worksheet name when --db is .xlsx (default: active)")
    parser.add_argument(
        "--db-no-header",
        action="store_true",
        help="Treat DB CSV/XLSX as headerless and read URLs from first column",
    )
    parser.add_argument(
        "--db-guid-column",
        default=None,
        help="Optional DB GUID column for matching GUID-like IDs in URLs",
    )
    parser.add_argument("--output", default=None, type=Path, help="Output CSV path")
    parser.add_argument(
        "--audit-output",
        default=None,
        type=Path,
        help="Output CSV path containing all URLs and how each was matched",
    )
    parser.add_argument(
        "--blocked",
        default=DEFAULT_BLOCKED_SOURCE,
        help="Blocked URL source: local path or URL (.txt, .csv, .xlsx)",
    )
    parser.add_argument("--blocked-url-column", default=None, help="Column name for blocked URLs when --blocked is CSV/XLSX")
    parser.add_argument("--blocked-sheet", default=None, help="Worksheet name when --blocked is .xlsx (default: active)")
    parser.add_argument("--blocked-no-header", action="store_true", help="Treat blocked file as headerless; read URLs from first column")

    parser.add_argument("--resolve-redirects", dest="resolve_redirects", action="store_true", default=True)
    parser.add_argument("--no-resolve-redirects", dest="resolve_redirects", action="store_false")
    parser.add_argument("--timeout", type=int, default=8, help="HTTP timeout seconds for redirect resolution")
    parser.add_argument("--workers", type=int, default=12, help="Concurrent workers for redirect checks")

    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)

    try:
        input_path = args.input if args.input else pick_input_file_interactively()
        input_stem = input_path.stem
        input_dir = input_path.parent

        output_path = args.output if args.output else input_dir / f"{input_stem}_truly_missing.csv"
        audit_output_path = args.audit_output if args.audit_output else input_dir / f"{input_stem}_match_audit.csv"

        cache_dir = Path(__file__).resolve().parent / ".cache"
        db_source_path = ensure_local_source(args.db, cache_dir, "ingested")
        blocked_source_path = ensure_local_source(args.blocked, cache_dir, "blocked") if args.blocked else None

        input_urls = read_urls(
            input_path,
            args.input_column,
            COMMON_URL_COLUMNS,
            sheet_name=args.input_sheet,
            no_header=args.input_no_header,
        )

        db_urls = read_urls(
            db_source_path,
            args.db_url_column,
            COMMON_URL_COLUMNS,
            sheet_name=args.db_sheet,
            no_header=args.db_no_header,
        )
        db_guids = read_db_guid_values(
            db_source_path,
            args.db_guid_column,
            sheet_name=args.db_sheet,
            no_header=args.db_no_header,
        )

        db_full, db_path, db_guid_in_urls = build_db_keys(db_urls)
        all_db_guids = db_guids | db_guid_in_urls

        blocked_full: set[str] | None = None
        blocked_path: set[str] | None = None
        blocked_count = 0
        if blocked_source_path:
            blocked_urls = read_urls(
                blocked_source_path,
                args.blocked_url_column,
                COMMON_URL_COLUMNS,
                sheet_name=args.blocked_sheet,
                no_header=args.blocked_no_header,
            )
            blocked_full, blocked_path, _ = build_db_keys(blocked_urls)
            blocked_count = len(blocked_urls)

        all_results = classify_all_inputs(
            input_urls=input_urls,
            db_full=db_full,
            db_path=db_path,
            db_guid=all_db_guids,
            resolve_redirects=args.resolve_redirects,
            timeout=args.timeout,
            workers=args.workers,
            blocked_full=blocked_full,
            blocked_path=blocked_path,
        )

        missing = [row for row in all_results if row.status == "missing"]
        not_testable = [row for row in all_results if row.status == "not-testable"]
        blocked = [row for row in all_results if row.status == "blocked"]

        write_results(output_path, missing)
        write_results(audit_output_path, all_results)

        found_count = sum(1 for row in all_results if row.status == "found")

        print(f"Input URLs: {len(input_urls)}")
        print(f"DB URLs: {len(db_urls)}")
        if args.blocked:
            print(f"Blocked URLs loaded: {blocked_count}")
        print(f"Detected DB GUIDs: {len(all_db_guids)}")
        print(f"Found URLs: {found_count}")
        print(f"Blocked URLs: {len(blocked)}")
        print(f"Not testable URLs: {len(not_testable)}")
        print(f"Truly missing URLs: {len(missing)}")
        print(f"Missing output written to: {output_path}")
        print(f"Audit output written to: {audit_output_path}")
        return 0
    except Exception as ex:  # noqa: BLE001
        print(f"Error: {type(ex).__name__}: {ex}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
